"""
run_backtest_dividende.py — Backtest ETF CGF Dividende BRVM
Méthodologie :
  - Univers    : toute la BRVM (titres avec historique dividende dans l'Excel)
  - Éligibilité: dividend yield > 0 sur ≥2 des 3 dernières années civiles
  - Score      : moyenne yield sur les années payées, plafonné à 20%/an (winsorisation)
  - Poids      : proportionnel au score, plafonné par ADV (liquidité)
  - Rebal      : annuel, 1er juillet (après saison dividendes BRVM avril–juin)
  - Type ETF   : Price Return — dividendes collectés et distribués aux porteurs
  - Spread     : tiéré par ADV (25/40/80/125/175 bps)
"""
import os, sys, json, warnings
from datetime import date, timedelta
from collections import defaultdict

import openpyxl

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

# ── Chemins ────────────────────────────────────────────────────────────────────
ROOT        = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(ROOT, "data")
EXCEL_PATH  = os.path.join(DATA_DIR, "BRVM_Consolidated_Kendall_updated.xlsx")
SIKA_PATH   = os.path.join(DATA_DIR, "sika_history.json")
NAV_PATH    = os.path.join(DATA_DIR, "nav_latest.json")
REBAL_PATH  = os.path.join(DATA_DIR, "rebal_detail.json")
LAUNCH_PATH = os.path.join(DATA_DIR, "launch_state.json")

# ── Paramètres ─────────────────────────────────────────────────────────────────
PAR_FCFA       = 100_000
N_PARTS        = 25_000
AUM_TARGET_M   = PAR_FCFA * N_PARTS / 1_000_000   # 2 500 M FCFA
MGMT_FEE_ANN   = 0.006
YIELD_CAP      = 0.20     # winsorisation : un yield individuel ne peut dépasser 20%
MIN_YEARS      = 2        # nb min d'années avec dividende sur les 3 de référence
MAX_TITRES     = 30       # max titres dans le panier (top 30 par score)
ADV_DAYS       = 20       # jours pour constituer/liquider la position
ETF_NAME       = "CGF ETF Dividende BRVM"
LAUNCH_DATE    = "2026-07-29"


# ── Spread tiéré ADV ───────────────────────────────────────────────────────────
def spread_one_way(adv_mfcfa: float) -> float:
    if adv_mfcfa >= 100: return 0.0025
    if adv_mfcfa >=  30: return 0.0040
    if adv_mfcfa >=  10: return 0.0080
    if adv_mfcfa >=   5: return 0.0125
    return 0.0175


# ── Chargement données dividendes depuis Excel ─────────────────────────────────
def load_div_yields() -> dict:
    """Retourne {ticker: {year_str: yield_float}} depuis feuille Dividend_Yield."""
    wb   = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws   = wb["📉 Dividend_Yield"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col    = {str(h): i for i, h in enumerate(header) if h is not None}

    years_available = [str(y) for y in range(2018, 2026) if str(y) in col]
    result = {}
    for row in rows[1:]:
        ticker = row[col["Symbol"]]
        if not ticker:
            continue
        yields = {}
        for y in years_available:
            val = row[col[y]]
            if val and val > 0:
                yields[y] = min(float(val), YIELD_CAP)
        result[ticker] = yields
    return result



# ── Calcul ADV sur les 90 derniers jours ──────────────────────────────────────
def compute_adv(sika_history: dict, ref_date: date) -> dict:
    adv = {}
    cutoff = (ref_date - timedelta(days=120)).isoformat()
    for tk, hist in sika_history.items():
        vols = []
        for d, v in hist.items():
            if d < cutoff:
                continue
            if isinstance(v, dict):
                vol   = v.get("volume", 0) or 0
                close = v.get("close", 0)  or 0
            else:
                vol, close = 0, 0
            if vol > 0 and close > 0:
                vols.append(vol * close / 1_000_000)
        adv[tk] = sum(vols) / len(vols) if vols else 0.0
    return adv


# ── Sélection + pondération à une date de rebalancement ───────────────────────
def compute_weights(ref_year: int, div_yields: dict, adv: dict) -> tuple:
    """
    ref_year : année de rebalancement (ex 2026 → utilise 2023,2024,2025)
    Retourne (w_etf, w_idx) :
      w_etf : {ticker: weight} après ADV cap (ce que l'ETF détient)
      w_idx : {ticker: weight} brut yield pur sans cap (indice de référence)
    """
    ref_years = [str(ref_year - 3), str(ref_year - 2), str(ref_year - 1)]

    scores = {}
    for tk, yields in div_yields.items():
        vals = [yields[y] for y in ref_years if y in yields and yields[y] > 0]
        if len(vals) >= MIN_YEARS:
            scores[tk] = sum(vals) / len(vals)

    if not scores:
        return {}, {}

    # Garder uniquement les MAX_TITRES meilleurs scores
    scores = dict(sorted(scores.items(), key=lambda x: -x[1])[:MAX_TITRES])

    # Poids indice : yield pur, sans cap (base 1.0)
    total_score = sum(scores.values())
    w_idx = {tk: s / total_score for tk, s in scores.items()}

    # Poids ETF : plafonné par ADV
    w_capped = {}
    for tk, w in w_idx.items():
        adv_tk = adv.get(tk, 0)
        w_max  = (adv_tk * ADV_DAYS) / AUM_TARGET_M if AUM_TARGET_M > 0 else 1.0
        w_capped[tk] = min(w, w_max) if w_max > 0 else w

    total = sum(w_capped.values()) or 1.0
    w_etf = {tk: w / total for tk, w in w_capped.items() if w > 0}

    return w_etf, w_idx


# ── Construction NAV (backtest) ────────────────────────────────────────────────
def build_nav(sika_history: dict, div_yields: dict, rebal_years: list) -> tuple:
    """
    Retourne (nav_etf_series, nav_idx_series, rebal_log, w_etf_history)
    nav_*_series : {date_str: float}  — base 100 au 1er juillet de rebal_years[0]
    """
    # Dates de rebalancement : 1er juillet de chaque année
    rebal_dates = []
    for yr in rebal_years:
        rd = date(yr, 7, 1)
        # Décaler au lundi si week-end
        while rd.weekday() >= 5:
            rd += timedelta(days=1)
        rebal_dates.append(rd.isoformat())

    start_date = rebal_dates[0]

    # Toutes les dates de trading dans sika_history depuis start_date
    all_dates_set = set()
    for hist in sika_history.values():
        for d in hist:
            if d >= start_date:
                all_dates_set.add(d)
    all_dates = sorted(d for d in all_dates_set
                       if date.fromisoformat(d).weekday() < 5)

    if not all_dates:
        return {}, {}, [], {}

    fee_daily = (1.0 - MGMT_FEE_ANN) ** (1.0 / 252.0)

    nav_etf = {}
    nav_idx = {}
    rebal_log = []
    w_etf_history = {}
    adv_at_rebal = {}

    curr_w_etf = {}
    curr_w_idx = {}
    nav_e = 100.0
    nav_i = 100.0

    prev_prices = {}

    for i, dt in enumerate(all_dates):
        d_obj    = date.fromisoformat(dt)
        prev_dt  = all_dates[i - 1] if i > 0 else None

        # Prix du jour
        prices = {}
        for tk, hist in sika_history.items():
            if dt in hist:
                v = hist[dt]
                p = v.get("close") if isinstance(v, dict) else v
                if p and p > 0:
                    prices[tk] = float(p)

        # ── Rebalancement ────────────────────────────────────────────────────
        if dt in rebal_dates:
            yr_idx    = rebal_dates.index(dt)
            ref_year  = rebal_years[yr_idx]
            adv       = compute_adv(sika_history, d_obj)
            adv_at_rebal[dt] = adv

            new_w_etf, new_w_idx = compute_weights(ref_year, div_yields, adv)

            # Spread de transaction
            all_tks = set(curr_w_etf) | set(new_w_etf)
            cost = sum(
                abs(new_w_etf.get(tk, 0) - curr_w_etf.get(tk, 0))
                * spread_one_way(adv.get(tk, 0))
                for tk in all_tks
            )
            turnover = sum(abs(new_w_etf.get(tk, 0) - curr_w_etf.get(tk, 0))
                           for tk in all_tks) / 2

            nav_e *= (1.0 - cost)

            # Log rebal — w_brvm30 = poids indice (yield pur, sans ADV cap)
            basket_log = []
            for tk, w in sorted(new_w_etf.items(), key=lambda x: -x[1]):
                adv_tk  = adv.get(tk, 0)
                y_ref   = [str(ref_year - 3), str(ref_year - 2), str(ref_year - 1)]
                avg_y   = sum(div_yields.get(tk, {}).get(y, 0) for y in y_ref
                              if div_yields.get(tk, {}).get(y, 0) > 0)
                n_y     = sum(1 for y in y_ref if div_yields.get(tk, {}).get(y, 0) > 0)
                avg_y   = avg_y / n_y if n_y else 0
                w_idx_tk = new_w_idx.get(tk, w)
                basket_log.append({
                    "ticker":        tk,
                    "w_etf":         round(w, 6),
                    "w_brvm30":      round(w_idx_tk, 6),   # poids indice pur
                    "avg_yield_pct": round(avg_y * 100, 2),
                    "adv_mfcfa":     round(adv_tk, 2),
                    "prix_rebal":    round(prices.get(tk, 0), 0),
                    "capped":        w < w_idx_tk - 1e-9,
                })
            rebal_log.append({
                "date":     dt,
                "ref_year": ref_year,
                "n_titres": len(new_w_etf),
                "turnover": round(turnover * 100, 2),
                "cout_pct": round(cost * 100, 4),
                "nav_etf":  round(nav_e, 4),
                "basket":   basket_log,
            })

            curr_w_etf = new_w_etf.copy()
            curr_w_idx = new_w_idx.copy()   # indice = yield pur, sans ADV cap
            prev_prices = {tk: prices.get(tk, prev_prices.get(tk, 1))
                           for tk in curr_w_etf}
            w_etf_history[dt] = curr_w_etf.copy()

        if not curr_w_etf:
            continue

        # ── Rendement journalier ─────────────────────────────────────────────
        ret_etf = 0.0
        ret_idx = 0.0
        for tk, w in curr_w_etf.items():
            p0 = prev_prices.get(tk)
            p1 = prices.get(tk)
            if p0 and p0 > 0 and p1 and p1 > 0:
                r = p1 / p0 - 1.0
                ret_etf += w * r
                ret_idx += w * r   # PR : dividendes exclus du NAV indice aussi

        # Frais de gestion quotidiens sur l'ETF
        nav_e = nav_e * (1.0 + ret_etf) * fee_daily
        nav_i = nav_i * (1.0 + ret_idx)

        nav_etf[dt] = nav_e
        nav_idx[dt] = nav_i

        # Mise à jour des poids (dérive mark-to-market)
        if prev_dt and prev_dt in nav_etf:
            new_ws = {}
            total = 0.0
            for tk, w in curr_w_etf.items():
                p0 = prev_prices.get(tk)
                p1 = prices.get(tk)
                if p0 and p0 > 0 and p1 and p1 > 0:
                    new_ws[tk] = w * (p1 / p0)
                    total += new_ws[tk]
                else:
                    new_ws[tk] = w
                    total += w
            curr_w_etf = {tk: w / total for tk, w in new_ws.items()} if total > 0 else curr_w_etf

        prev_prices = {tk: prices.get(tk, prev_prices.get(tk, 1))
                       for tk in curr_w_etf}

    return nav_etf, nav_idx, rebal_log, w_etf_history


# ── Métriques de performance ──────────────────────────────────────────────────
def compute_metrics(nav_series: dict, label: str) -> dict:
    dates = sorted(nav_series.keys())
    if len(dates) < 2:
        return {}
    vals  = [nav_series[d] for d in dates]
    n_days = (date.fromisoformat(dates[-1]) - date.fromisoformat(dates[0])).days
    n_years = n_days / 365.25

    perf_total = vals[-1] / vals[0] - 1.0
    perf_ann   = (1.0 + perf_total) ** (1.0 / n_years) - 1.0 if n_years > 0 else 0

    import math
    rets = [(vals[i] - vals[i-1]) / vals[i-1] for i in range(1, len(vals))]
    vol  = (sum(r**2 for r in rets) / len(rets) - (sum(rets)/len(rets))**2) ** 0.5
    vol_ann = vol * math.sqrt(252)

    peak = vals[0]
    max_dd = 0.0
    for v in vals:
        if v > peak:
            peak = v
        dd = (v - peak) / peak
        if dd < max_dd:
            max_dd = dd

    print(f"\n{label}:")
    print(f"  Perf totale  : {perf_total*100:+.2f}%")
    print(f"  Perf ann.    : {perf_ann*100:+.2f}%/an")
    print(f"  Vol ann.     : {vol_ann*100:.2f}%")
    print(f"  Max Drawdown : {max_dd*100:.2f}%")
    print(f"  Sharpe (rf=0): {(perf_ann/vol_ann):.2f}" if vol_ann > 0 else "  Sharpe       : —")

    return {
        "perf_total_pct": round(perf_total * 100, 2),
        "perf_ann_pct":   round(perf_ann   * 100, 2),
        "vol_ann_pct":    round(vol_ann    * 100, 2),
        "max_drawdown_pct": round(max_dd   * 100, 2),
        "sharpe":         round(perf_ann / vol_ann, 2) if vol_ann > 0 else None,
        "start_date":     dates[0],
        "end_date":       dates[-1],
    }


# ── Initialisation nav_latest.json ────────────────────────────────────────────
def init_nav_latest(w_etf_history, sika_history, rebal_log, nav_etf, nav_idx,
                    div_yields, metrics_etf, metrics_idx):
    last_rebal_date = sorted(w_etf_history.keys())[-1]
    curr_w          = w_etf_history[last_rebal_date]

    # Dernier rebal log pour le basket
    last_log = next((r for r in reversed(rebal_log)
                     if r["date"] == last_rebal_date), {})
    basket_map = {b["ticker"]: b for b in last_log.get("basket", [])}

    # Prix les plus récents
    latest_prices = {}
    for tk, hist in sika_history.items():
        if hist:
            ld = max(hist.keys())
            v  = hist[ld]
            p  = v.get("close") if isinstance(v, dict) else v
            if p:
                latest_prices[tk] = float(p)

    # Poids flottants depuis dernier rebal
    floated = {}
    for tk, w0 in curr_w.items():
        p_rebal = basket_map.get(tk, {}).get("prix_rebal", 0)
        p_now   = latest_prices.get(tk, 0)
        if p_rebal and p_rebal > 0 and p_now > 0:
            floated[tk] = w0 * (p_now / p_rebal)
        else:
            floated[tk] = w0
    total = sum(floated.values()) or 1.0
    floated = {tk: w / total for tk, w in floated.items()}

    basket = []
    for tk, w in sorted(floated.items(), key=lambda x: -x[1]):
        b = basket_map.get(tk, {})
        basket.append({
            "ticker":        tk,
            "poids_pct":     round(w * 100, 4),
            "w_brvm30":      round(b.get("w_brvm30", w), 6),   # poids indice
            "avg_yield_pct": b.get("avg_yield_pct", 0),
            "adv_mfcfa":     b.get("adv_mfcfa", 0),
            "dernier_prix":  round(latest_prices.get(tk, b.get("prix_rebal", 0)), 0),
            "prix_rebal":    b.get("prix_rebal", 0),
            "capped":        b.get("capped", False),
        })

    # nav_series (backtest complet, PR)
    all_dates = sorted(nav_etf.keys())
    nav_series = [[d, round(nav_etf[d], 4)] for d in all_dates
                  if date.fromisoformat(d).weekday() < 5]

    nav_indice_now = nav_etf[max(nav_etf.keys())] if nav_etf else 100.0

    # Préserver les données live si l'ETF est déjà lancé
    existing_nl = {}
    if os.path.exists(NAV_PATH):
        with open(NAV_PATH, encoding="utf-8") as f:
            existing_nl = json.load(f)
    already_launched = existing_nl.get("launched", False)

    nl = {
        "etf_name":            ETF_NAME,
        "calc_date":           existing_nl.get("calc_date") if already_launched else None,
        "launched":            already_launched,
        "nav_indice":          existing_nl.get("nav_indice", round(nav_indice_now, 4)) if already_launched else round(nav_indice_now, 4),
        "vl_par_part_fcfa":    existing_nl.get("vl_par_part_fcfa", PAR_FCFA) if already_launched else PAR_FCFA,
        "aum_mfcfa":           existing_nl.get("aum_mfcfa", round(PAR_FCFA * N_PARTS / 1_000_000, 1)) if already_launched else round(PAR_FCFA * N_PARTS / 1_000_000, 1),
        "n_parts":             N_PARTS,
        "perf_since_launch":   existing_nl.get("perf_since_launch", 0.0) if already_launched else 0.0,
        "change_day_pct":      existing_nl.get("change_day_pct") if already_launched else None,
        "last_rebal_date":     LAUNCH_DATE,
        "nav_live_series":     existing_nl.get("nav_live_series", []) if already_launched else [],
        "nav_indice_reference": existing_nl.get("nav_indice_reference") if already_launched else (round(nav_idx[max(nav_idx.keys())], 4) if nav_idx else None),
        "nav_series":          nav_series,
        "perf_backtest_total": metrics_etf.get("perf_total_pct"),
        "perf_backtest_ann":   metrics_etf.get("perf_ann_pct"),
        "vol_ann_pct":         metrics_etf.get("vol_ann_pct"),
        "max_drawdown_pct":    metrics_etf.get("max_drawdown_pct"),
        "sharpe":              metrics_etf.get("sharpe"),
        "perf_indice_total":    metrics_idx.get("perf_total_pct"),
        "perf_indice_ann":      metrics_idx.get("perf_ann_pct"),
        "basket":               basket,
        "n_basket":            len(basket),
        "methodology":         "price_return",
        "rebal_frequency":     "annual_july",
        "yield_cap_pct":       YIELD_CAP * 100,
        "min_years_dividend":  MIN_YEARS,
    }
    with open(NAV_PATH, "w", encoding="utf-8") as f:
        json.dump(nl, f, ensure_ascii=False, indent=2)
    print(f"\nnav_latest.json écrit : {len(basket)} titres, nav_indice={nav_indice_now:.4f}")


# ── Validation results (format attendu par generate_dashboard_data.py) ────────
def save_validation_results(nav_etf, nav_idx, m_etf, m_idx, rebal_log):
    import math
    all_dates = sorted(nav_etf.keys())
    nav_etf_list = [[d, round(nav_etf[d], 4)] for d in all_dates]
    nav_idx_list = [[d, round(nav_idx[d], 4)] for d in sorted(nav_idx.keys())]

    # Tracking error et tracking difference
    common = sorted(set(nav_etf) & set(nav_idx))
    diff_rets = []
    for i in range(1, len(common)):
        r_e = nav_etf[common[i]] / nav_etf[common[i-1]] - 1
        r_i = nav_idx[common[i]] / nav_idx[common[i-1]] - 1
        diff_rets.append(r_e - r_i)
    mean_d = sum(diff_rets) / len(diff_rets) if diff_rets else 0
    var_d  = sum((r - mean_d)**2 for r in diff_rets) / len(diff_rets) if diff_rets else 0
    te_ann = math.sqrt(var_d * 252) * 100

    perf_etf   = m_etf.get("perf_total_pct", 0) / 100
    perf_bench = m_idx.get("perf_total_pct", 0) / 100
    n_years    = m_etf.get("n_days", 252 * 3) / 252 if "n_days" in m_etf else (
        (date.fromisoformat(all_dates[-1]) - date.fromisoformat(all_dates[0])).days / 365.25
    )
    td_ann = ((1 + perf_etf) ** (1 / n_years) - 1 - ((1 + perf_bench) ** (1 / n_years) - 1)) * 100 if n_years > 0 else 0

    # Perf par année
    perf_by_year = {}
    all_yr = sorted({d[:4] for d in nav_etf})
    for yr in all_yr:
        dates_yr = [d for d in sorted(nav_etf) if d.startswith(yr)]
        if len(dates_yr) < 2:
            continue
        r_etf = nav_etf[dates_yr[-1]] / nav_etf[dates_yr[0]] - 1
        r_idx = nav_idx.get(dates_yr[-1], nav_idx.get(dates_yr[0], nav_idx[min(nav_idx.keys())])) / \
                nav_idx.get(dates_yr[0], nav_idx[min(nav_idx.keys())]) - 1
        perf_by_year[yr] = {
            "etf_pct":    round(r_etf * 100, 2),
            "indice_pct": round(r_idx * 100, 2),
            "td_pct":     round((r_etf - r_idx) * 100, 2),
        }

    total_turnover = sum(r.get("turnover", 0) / 100 for r in rebal_log)
    avg_turnover   = total_turnover / len(rebal_log) * 100 if rebal_log else 0

    vr = {
        "generated_at":               date.today().isoformat(),
        "backtest_start":             all_dates[0] if all_dates else "",
        "backtest_end":               all_dates[-1] if all_dates else "",
        "metrics_etf":                m_etf,
        "metrics_indice":             m_idx,
        "tracking_difference_ann_pct": round(td_ann, 2),
        "tracking_error_ann_pct":      round(te_ann, 2),
        "total_turnover_rebal":        round(total_turnover, 4),
        "avg_turnover_per_rebal_pct":  round(avg_turnover, 2),
        "perf_by_year":                perf_by_year,
        "nav_etf_series":              nav_etf_list[-500:],
        "nav_index_series":            nav_idx_list[-500:],
    }
    vr_path = os.path.join(DATA_DIR, "validation_results.json")
    with open(vr_path, "w", encoding="utf-8") as f:
        json.dump(vr, f, ensure_ascii=False, indent=2)
    print(f"validation_results.json écrit ({len(nav_etf_list)} points ETF)")


# ── Rebal detail ──────────────────────────────────────────────────────────────
def save_rebal_detail(rebal_log):
    rd = {"etf": ETF_NAME, "rebalancings": rebal_log}
    with open(REBAL_PATH, "w", encoding="utf-8") as f:
        json.dump(rd, f, ensure_ascii=False, indent=2)
    print(f"rebal_detail.json écrit : {len(rebal_log)} rebalancements")


# ── Complétion dividend_history.json depuis Excel (tickers absents de Sika) ───
def complete_dividend_history_from_excel():
    dh_path = os.path.join(DATA_DIR, "dividend_history.json")
    if not os.path.exists(dh_path):
        return
    with open(dh_path, encoding="utf-8") as f:
        dh = json.load(f)
    hist = dh.get("history", {})

    try:
        wb  = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws  = next(wb[s] for s in wb.sheetnames if "Dividende" in s)
    except Exception as e:
        print(f"[WARN] Excel dividendes non lisible : {e}")
        return

    rows   = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col    = {str(h): i for i, h in enumerate(header) if h is not None}
    years  = [str(y) for y in range(2018, 2027) if str(y) in col]

    n_added = 0
    for row in rows[1:]:
        tk = row[col.get("Symbol", 0)]
        if not tk:
            continue
        for yr in years:
            val = row[col[yr]] if yr in col else None
            if not val or float(val) <= 0:
                continue
            if tk not in hist:
                hist[tk] = {}
            if not hist[tk].get(yr):  # ne pas écraser les données Sika
                hist[tk][yr] = float(val)
                n_added += 1

    dh["history"] = hist
    with open(dh_path, "w", encoding="utf-8") as f:
        json.dump(dh, f, ensure_ascii=False, indent=2)
    print(f"[OK] {n_added} dividendes ajoutés depuis Excel dans dividend_history.json")


# ── Fichiers live vides ────────────────────────────────────────────────────────
def init_live_files():
    for path, content in [
        (os.path.join(DATA_DIR, "intraday_nav.json"),          {"date": None, "snapshots": []}),
        (os.path.join(DATA_DIR, "nav_intraday_history.json"),  {}),
        (os.path.join(DATA_DIR, "dividend_log.json"),          {}),
        (os.path.join(DATA_DIR, "dividend_history.json"),      {"distributions": []}),
    ]:
        if not os.path.exists(path):
            with open(path, "w", encoding="utf-8") as f:
                json.dump(content, f, ensure_ascii=False, indent=2)

    # Ne pas écraser launch_state.json si déjà créé par le lancement réel
    if not os.path.exists(LAUNCH_PATH):
        ls = {
            "etf_name":            ETF_NAME,
            "launch_date":         LAUNCH_DATE,
            "nav_index_at_launch": None,
            "par_fcfa":            PAR_FCFA,
            "n_parts":             N_PARTS,
        }
        with open(LAUNCH_PATH, "w", encoding="utf-8") as f:
            json.dump(ls, f, ensure_ascii=False, indent=2)
        print("launch_state.json et fichiers live initialisés")
    else:
        print("launch_state.json déjà existant — conservé")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print(f"Backtest {ETF_NAME}")
    print("=" * 60)

    print("\n[1] Chargement données...")
    div_yields   = load_div_yields()
    sika_history = json.load(open(SIKA_PATH, encoding="utf-8"))
    print(f"    Dividendes : {len(div_yields)} tickers")
    print(f"    Prix       : {len(sika_history)} tickers")

    # Années de rebalancement disponibles : 2021 → 2026
    # Besoin de 3 ans d'historique avant le 1er rebal → premier rebal 2021 (réf 2018-2020)
    rebal_years = [2021, 2022, 2023, 2024, 2025, 2026]

    print("\n[2] Backtest en cours...")
    nav_etf, nav_idx, rebal_log, w_etf_history = build_nav(
        sika_history, div_yields, rebal_years
    )

    print(f"    {len(nav_etf)} jours de trading simulés")
    print(f"    {len(rebal_log)} rebalancements")

    # Afficher composition du dernier rebal
    if rebal_log:
        last = rebal_log[-1]
        print(f"\n    Dernier rebal {last['date']} (réf {last['ref_year']}) — "
              f"{last['n_titres']} titres, turnover {last['turnover']}%, coût {last['cout_pct']}%")
        for b in last["basket"][:10]:
            print(f"      {b['ticker']:<8} {b['w_etf']*100:6.2f}%  yield_moy={b['avg_yield_pct']:.1f}%  "
                  f"ADV={b['adv_mfcfa']:.0f}M  {'[CAP]' if b['capped'] else ''}")

    print("\n[3] Métriques...")
    m_etf = compute_metrics(nav_etf, "ETF Dividende (PR, frais+spread)")
    m_idx = compute_metrics(nav_idx, "Indice Dividende (PR, sans frais)")

    print("\n[4] Écriture fichiers...")
    init_live_files()
    complete_dividend_history_from_excel()
    init_nav_latest(w_etf_history, sika_history, rebal_log, nav_etf, nav_idx,
                    div_yields, m_etf, m_idx)
    save_rebal_detail(rebal_log)
    save_validation_results(nav_etf, nav_idx, m_etf, m_idx, rebal_log)

    print("\n[5] Génération dashboard_data.json...")
    import subprocess
    dash_script = os.path.join(os.path.dirname(__file__), "generate_dashboard_data.py")
    subprocess.run([sys.executable, dash_script], check=True)

    print("\n✓ Backtest terminé.")


if __name__ == "__main__":
    main()
